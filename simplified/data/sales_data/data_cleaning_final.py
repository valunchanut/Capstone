import os
import pandas as pd
import openpyxl
from datetime import datetime

def clean_sales_data(sales_mix_directory='sales_mix'):
    date_product_data = {}
    
    for file_name in os.listdir(sales_mix_directory):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(sales_mix_directory, file_name)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Find the header row. Assuming 'Name' and 'Quantity Sold' could be in any column.
            header_row_idx = None
            for rowIndex, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if 'Name' in row and 'Quantity Sold' in row:
                    header_row_idx = rowIndex
                    headers = [cell for cell in row if cell is not None]
                    name_idx = headers.index('Name')
                    quantity_sold_idx = headers.index('Quantity Sold')
                    break

            if header_row_idx is None:
                raise ValueError(f"'Name' and 'Quantity Sold' columns not found in the file: {file_name}")

            # Parse the date from the filename
            date_part = ' '.join(file_name.split(' ')[2:4]).replace(' - Copy', '').strip()
            date_product_data[date_part] = date_product_data.get(date_part, {})

            for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                product_name = row[name_idx]
                quantity_sold = row[quantity_sold_idx] or 0

                names_to_remove = [...]

                if any(pattern.lower() in product_name.lower() for pattern in ['allergy', 'combo', '$', 'add', 'No']):
                    continue
                if product_name in names_to_remove:
                    continue

                date_product_data[date_part][product_name] = date_product_data[date_part].get(product_name, 0) + quantity_sold

    product_sales_df = pd.DataFrame.from_dict(date_product_data, orient='columns')
    product_sales_df = product_sales_df.drop(index=names_to_remove, errors='ignore').transpose()

    # Function to adjust the year based on month
    def adjust_year(date_str):
        date_obj = datetime.strptime(date_str + " 2023", "%b %d %Y")  # Initial assumption of year
        # Adjust to 2024 if the month is January
        if date_obj.month == 1:
            date_obj = date_obj.replace(year=2024)
        return date_obj

    # Adjust the index for each row in the DataFrame
    adjusted_dates = [adjust_year(date) for date in product_sales_df.index]
    product_sales_df.index = pd.to_datetime(adjusted_dates)
    product_sales_df.sort_index(inplace=True)

    # Assume filling missing values with 0 and ensuring integer data types
    product_sales_df = product_sales_df.fillna(0).astype(int)

    return product_sales_df

if __name__ == '__main__':
    cleaned_data = clean_sales_data()
    output_csv_path = 'combined_sales_data.csv'
    cleaned_data.to_csv(output_csv_path, index_label='Date')
    print(cleaned_data.head())
